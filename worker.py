import openpyxl
import logging
from datetime import datetime


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelHandler:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb[sheet_name]

    def get_value(self, row, column):
        return self.ws.cell(row=row, column=column).value

    def set_value(self, row, column, value):
        self.ws.cell(row=row, column=column).value = value
        logger.info(f'Set value {value} in row {row}, column {column}')

    def save(self):
        self.wb.save(self.file_path)
        logger.info(f'Saved file {self.file_path}')


class ExchangeRateFinder:
    def __init__(self, excel_handler):
        self.excel_handler = excel_handler

    def get_rates_by_date(self, date):
        row = 2
        current_value = self.excel_handler.get_value(row, 2)
        current_date = None

        while current_value:
            if isinstance(current_value, str):
                current_date = current_value
            elif isinstance(current_value, datetime):
                current_date = str(current_value.strftime("%Y-%m-%d"))

            if current_date == date.strftime("%Y-%m-%d"):
                usd = self.excel_handler.get_value(row, 3)
                usd_deferred = self.excel_handler.get_value(row, 4)
                eur = self.excel_handler.get_value(row, 5)
                eur_deferred = self.excel_handler.get_value(row, 6)
                return {"usd": usd, "usd_deferred": usd_deferred, "eur": eur, "eur_deferred": eur_deferred}
            row += 1
            current_value = self.excel_handler.get_value(row, 2)

        return None

class Logger:
    @staticmethod
    def log_rates(formatted_today, rates):
        if rates:
            logger.info(f"Exchange rates for {formatted_today}:")
            logger.info(f"USD: {rates['usd']}")
            logger.info(f"USD Deferred: {rates['usd_deferred']}")
            logger.info(f"EUR: {rates['eur']}")
            logger.info(f"EUR Deferred: {rates['eur_deferred']}")
        else:
            logger.warning(f"No exchange rates found for {formatted_today}.")

class ExchangeRateRecorder:
    def __init__(self, excel_handler):
        self.excel_handler = excel_handler

    def record_rates(self, date, rates):
        self.excel_handler.set_value(2, 2, date)  # Cell R2C2
        self.excel_handler.set_value(3, 4, rates['usd'])  # Cell R3C4
        self.excel_handler.set_value(3, 5, rates['usd_deferred'])  # Cell R3C5
        self.excel_handler.set_value(4, 4, rates['eur'])  # Cell R4C4
        self.excel_handler.set_value(4, 5, rates['eur_deferred'])  # Cell R4C5
        logger.info(f'Recorded exchange rates for {date}: {rates}')



class PriceCalculator:
    def __init__(self, excel_handler, rates):
        self.excel_handler = excel_handler
        self.rates = rates

    def calculate_prices_euro(self, row, price):
        eur_price = price * self.rates['eur']
        eur_deferred_price = price * self.rates['eur_deferred']

        logger.info(f'Calculated EUR prices for row {row}: {eur_price}, {eur_deferred_price}')

        self.excel_handler.set_value(row, 12, eur_price)  # Column L
        self.excel_handler.set_value(row, 13, eur_deferred_price)  # Column M

    def calculate_prices_dollar(self, row, price):
        usd_price = price * self.rates['usd']
        usd_deferred_price = price * self.rates['usd_deferred']

        logger.info(f'Calculated USD prices for row {row}: {usd_price}, {usd_deferred_price}')

        self.excel_handler.set_value(row, 12, usd_price)  # Column L
        self.excel_handler.set_value(row, 13, usd_deferred_price)  # Column M

    def calculate_prices(self):
        logger.info('Start calculating prices')
        row = 7
        cell_value = self.excel_handler.get_value(row, 2)  # Column B

        while cell_value:
            currency = self.excel_handler.get_value(row, 11)  # Column K
            price = self.excel_handler.get_value(row, 10)  # Column J

            if currency == 'EVRO':
                self.calculate_prices_euro(row, price)
            elif currency == 'DOLLAR':
                self.calculate_prices_dollar(row, price)

            row += 1
            cell_value = self.excel_handler.get_value(row, 2)  # Column B

def main():
    file_path = "Pricess2.xlsx"
    sheet_name1 = "Kurs"
    sheet_name2 = "Price2"
    today = datetime.now().date()
    formatted_today = today.strftime("%Y-%m-%d")

    excel_handler_1 = ExcelHandler(file_path, sheet_name1)
    excel_handler_2 = ExcelHandler(file_path, sheet_name2)
    exchange_rate_finder = ExchangeRateFinder(excel_handler_1)
    rates = exchange_rate_finder.get_rates_by_date(today)

    if rates:
        logger.info(f"Exchange rates for {formatted_today}:")
        logger.info(f"USD: {rates['usd']}")
        logger.info(f"USD Deferred: {rates['usd_deferred']}")
        logger.info(f"EUR: {rates['eur']}")
        logger.info(f"EUR Deferred: {rates['eur_deferred']}")

        exchange_rate_recorder = ExchangeRateRecorder(excel_handler_2)
        exchange_rate_recorder.record_rates(formatted_today, rates)

        price_calculator = PriceCalculator(excel_handler_2, rates)
        price_calculator.calculate_prices()

        excel_handler_2.save()
        logger.info('Finished all operations and saved the Excel file.')

    else:
        logger.warning(f"No exchange rates found for {formatted_today}.")

if __name__ == "__main__":
    main()


